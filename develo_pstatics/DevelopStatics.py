# coding:utf-8
import datetime
import pymysql
import openpyxl
from Tools.demo.mcast import receiver
from openpyxl.styles import Font
from send_email import EmailSender


class ZenTaoStatics(object):
    wb = None
    lmld = ''
    tmpd = ''

    def __init__(self):
        self.lmld = self.get_lmld()
        self.tmpd = self.get_tmpd()

    @staticmethod
    def get_lmld():
        """
        获取上个月最后一天的日期
        :return: string
        """
        last_month_last_day = datetime.date(datetime.date.today().year, datetime.date.today().month,
                                            1) - datetime.timedelta(1)
        return last_month_last_day.strftime('%Y-%m-%d')

    @staticmethod
    def get_tmpd():
        """
        获取获得本月中的倒数第二天
        :return: string
        """
        next_month = datetime.date.today().replace(day=28) + datetime.timedelta(days=4)  # this will never fail
        this_month_penultimate_day = next_month - datetime.timedelta(days=2)
        return this_month_penultimate_day.strftime('%Y-%m-%d')

    def getDbConnect(self):
        db = pymysql.connect(host="10.168.7.66", user="root", password="HZ123tech", db="zentao", port=3306,
                             charset='utf8')
        return db

    def query(self, sql):
        try:
            db = self.getDbConnect()
            cur = db.cursor()
            effectrow = cur.execute(sql)
            if effectrow == 0:
                return []
            return cur.fetchall()
        except Exception as e:
            print("数据库执行异常:%s" % e)
        finally:
            db.close()

    def get_developer_bugs(self):
        """查询开发人员本月bug汇总"""
        sql_query = """SELECT
                            u.realname "开发人员姓名",
                            b.id "Bug ID",
                            b.title "Bug标题",
                            b.severity "严重等级",
                            b.openedDate "创建日期",
                            b.resolvedDate "解决日期"
                        FROM
                            zt_bug b
                            LEFT JOIN zt_user u ON b.assignedTo = u.account 
                            OR b.resolvedBy = u.account 
                            LEFT JOIN zt_dept d ON u.dept = d.id
                        WHERE
                            b.deleted = '0' 
                            AND instr( d.path, '2' ) > 0 
                            AND b.openedDate BETWEEN '%s' AND '%s'
                            ORDER BY u.realname asc""" % (self.lmld, self.tmpd)
        ws = self.wb.create_sheet('{}-{}开发人员bug汇总'.format(self.lmld, self.tmpd))
        title = ['开发人员姓名', 'Bug ID', 'Bug标题', '严重等级', '创建日期', '解决日期']
        ws.append(title)
        statics_result = self.query(sql_query)
        row_count = 2
        for row in statics_result:
            ws.append(row)
            ws.cell(row=row_count, column=2).hyperlink = "http://10.168.7.66/zentao/bug-view-%s.html" % row[1]
            ws.cell(row=row_count, column=2).font = Font(color="0000FF", underline='single')
            row_count += 1

    def get_tester_bugs(self):
        """查询测试人员本月bug汇总"""
        sql_query = """SELECT
                            u.realname "测试姓名",
                            proj.`name` "项目名称",
                            b.id "Bug ID",
                            b.title "Bug标题",
                            b.severity "严重等级",
                            b.openedDate "创建日期"
                        FROM
                            zt_bug b
                            LEFT JOIN zt_user u ON b.openedBy = u.account 
                            LEFT JOIN zt_dept d ON u.dept = d.id
                            LEFT JOIN zt_project proj ON b.project = proj.id 
                        WHERE
                            b.deleted = '0' 
                            AND instr( d.path, '3' ) > 0 
                            AND b.openedDate BETWEEN '%s' AND '%s'
                            ORDER BY u.realname, proj.`name` asc""" % (self.lmld, self.tmpd)
        ws = self.wb.create_sheet('{}-{}测试人员bug汇总'.format(self.lmld, self.tmpd))
        title = ['测试姓名', '项目名称', 'Bug ID', 'Bug标题', '严重等级', '创建日期']
        ws.append(title)
        statics_result = self.query(sql_query)
        row_count = 2
        for row in statics_result:
            ws.append(row)
            ws.cell(row=row_count, column=3).hyperlink = "http://10.168.7.66/zentao/bug-view-%s.html" % row[2]
            ws.cell(row=row_count, column=3).font = Font(color="0000FF", underline='single')
            row_count += 1

    def get_developer_bugs_statics(self):
        """查询开发人员本月产生的bug数量"""
        sql_query = """SELECT
                        d2.`name` 所属产品线,
                            d1.`name` 所属小组,
                            proj.NAME 项目名,
                            u.realname 姓名,
                            count(*) "本月产生bug数量",
                            sum( CASE WHEN b.severity = 1 THEN 1 ELSE 0 END ) "致命bug数量",
                            GROUP_CONCAT( CASE WHEN b.severity = 1 THEN b.id ELSE NULL END ) "致命bug ID",
                            sum( CASE WHEN b.severity = 2 THEN 1 ELSE 0 END ) "严重bug数量",
                            GROUP_CONCAT( CASE WHEN b.severity = 2 THEN b.id ELSE NULL END ) "严重bug ID",
                            sum( CASE WHEN b.severity = 3 THEN 1 ELSE 0 END ) "普通bug数量",
                            GROUP_CONCAT( CASE WHEN b.severity = 3 THEN b.id ELSE NULL END ) "普通bug ID",
                            sum( CASE WHEN b.severity = 4 THEN 1 ELSE 0 END ) "优化bug数量",
                            GROUP_CONCAT( CASE WHEN b.severity = 4 THEN b.id ELSE NULL END ) "优化bug ID"                            
                        FROM
                            zt_bug b
                            LEFT JOIN zt_user u ON b.assignedTo = u.account 
                            OR b.resolvedBy = u.account
                            LEFT JOIN zt_dept d1 ON u.dept = d1.id
                            LEFT JOIN zt_dept d2 ON d1.parent = d2.id
                            LEFT JOIN zt_project proj ON b.project = proj.id 
                        WHERE
                            b.deleted = '0'
                            AND instr( d1.path, '2' ) > 0 
                            AND b.openedDate BETWEEN '%s' AND '%s'
                        GROUP BY
                            u.realname 
                        ORDER BY
                            所属产品线,
                            所属小组,
                            项目名,
                            "本月产生bug数量" DESC""" % (self.lmld, self.tmpd)
        sheet = []
        title = ['所属产品线', '所属小组', '项目名', '开发姓名', '本月产生bug数量', '致命bug数量', '致命bug ID', '严重bug数量', '严重bug ID', '普通bug数量',
                 '普通bug ID', '优化bug数量', '优化bug ID']
        sheet.append(title)
        statics_result = self.query(sql_query)
        for row in statics_result:
            sheet.append(row)
        self.new_sheet('{}-{}开发人员产生bug数量'.format(self.lmld, self.tmpd), sheet)

    def get_tester_bugs_statics(self):
        """查询测试人员本月产生的bug数量"""
        sql_query = """SELECT
                            u.realname 测试姓名,
                            proj.NAME 项目名,
                            count(*) bug_count,
                            sum( CASE WHEN b.severity = 1 THEN 1 ELSE 0 END ) "致命bug数量",
                            GROUP_CONCAT( CASE WHEN b.severity = 1 THEN b.id ELSE NULL END ) "致命bug ID",
                            sum( CASE WHEN b.severity = 2 THEN 1 ELSE 0 END ) "严重bug数量",
                            GROUP_CONCAT( CASE WHEN b.severity = 2 THEN b.id ELSE NULL END ) "严重bug ID",
                            sum( CASE WHEN b.severity = 3 THEN 1 ELSE 0 END ) "普通bug数量",
                            GROUP_CONCAT( CASE WHEN b.severity = 3 THEN b.id ELSE NULL END ) "普通bug ID",
                            sum( CASE WHEN b.severity = 4 THEN 1 ELSE 0 END ) "优化bug数量",
                            GROUP_CONCAT( CASE WHEN b.severity = 4 THEN b.id ELSE NULL END ) "优化bug ID" 
                        FROM
                            zt_bug b
                            LEFT JOIN zt_user u ON b.openedBy = u.account 
                            LEFT JOIN zt_dept d ON u.dept = d.id
                            LEFT JOIN zt_project proj ON b.project = proj.id 
                        WHERE
                            b.deleted = '0'
                            AND instr( d.path, '3' ) > 0 
                            AND b.openedDate BETWEEN '%s' AND '%s'
                        GROUP BY
                        u.realname ,
                        proj.`name`
                        ORDER BY
                        u.realname,
                        bug_count DESC""" % (self.lmld, self.tmpd)
        sheet = []
        title = ['测试姓名', '项目名称', '本月提交bug数量', '致命bug数量', '致命bug ID', '严重bug数量', '严重bug ID', '普通bug数量', '普通bug ID',
                 '优化bug数量', '优化bug ID']
        sheet.append(title)
        statics_result = self.query(sql_query)
        for row in statics_result:
            sheet.append(row)
        self.new_sheet('{}-{}测试人员产生bug数量'.format(self.lmld, self.tmpd), sheet)

    def new_sheet(self, sheet_name, sheet_content):
        ws = self.wb.create_sheet(sheet_name)
        for row in sheet_content:
            ws.append(row)

    def save_workbook(self, file_name):
        self.wb.remove(self.wb.get_sheet_by_name('Sheet'))
        self.wb.save(file_name)

    def static_developer_bug_2xls(self):
        self.wb = openpyxl.workbook.Workbook()
        self.get_developer_bugs_statics()
        self.get_developer_bugs()
        self.save_workbook('本月开发人员Bug统计表.xlsx')

    def static_tester_bug_2xls(self):
        self.wb = openpyxl.workbook.Workbook()
        self.get_tester_bugs_statics()
        self.get_tester_bugs()
        self.save_workbook('本月测试人员Bug统计表.xlsx')


if __name__ == '__main__':
    # zt = ZenTaoStatics()
    # zt.static_developer_bug_2xls()
    # zt.static_tester_bug_2xls()
    es = EmailSender()
    es.set_email_subject("本月开发人员bug统计信息")
    es.set_email_body("大家好！\r\n    这是本月的开发人员Bug统计信息！请查收！")
    es.set_email_attach('本月开发人员Bug统计表.xlsx')
    receiver_list = ['dengjianghua']
    acc_list = ['wenxinrong']
    receiver = ['{}@hztech.cn'.format(x) for x in receiver_list]
    acc = ['{}@hztech.cn'.format(x) for x in acc_list]
    es.send_email(receiver=receiver, acc=acc)

    es = EmailSender()
    es.set_email_subject("本月测试人员bug统计信息")
    es.set_email_body("大家好！\r\n    这是本月的测试人员Bug统计信息！请查收！")
    es.set_email_attach('本月测试人员Bug统计表.xlsx')
    receiver_list = ['dengjianghua']
    acc_list = ['wenxinrong']
    receiver = ['{}@hztech.cn'.format(x) for x in receiver_list]
    acc = ['{}@hztech.cn'.format(x) for x in acc_list]
    es.send_email(receiver=receiver, acc=acc)
